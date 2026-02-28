import openpyxl
import os

# Read email_dep.xlsx for department lookup
dep_wb = openpyxl.load_workbook('public/email_dep.xlsx')
dep_ws = dep_wb.active

dept_lookup = {}
for row in dep_ws.iter_rows(min_row=2, values_only=True):
    email = row[1]
    dept = row[4]
    if email and dept:
        dept_lookup[str(email).strip().lower()] = str(dept).strip()

print(f"Department entries: {len(dept_lookup)}")
dep_wb.close()

# Read emails.xlsx
src_wb = openpyxl.load_workbook('public/emails.xlsx')
src_ws = src_wb.active

headers = [c.value for c in src_ws[1]]
email_col = next(i for i, h in enumerate(headers) if h and 'email' in str(h).lower() and 'address' in str(h).lower())

# Create new workbook
out_wb = openpyxl.Workbook()
out_ws = out_wb.active
out_ws.title = "Email Users"

# Write headers
new_headers = headers + ['Department']
for ci, h in enumerate(new_headers, 1):
    cell = out_ws.cell(row=1, column=ci, value=h)
    cell.font = openpyxl.styles.Font(bold=True)

# Write data
matched = 0
row_num = 2
for row in src_ws.iter_rows(min_row=2, values_only=True):
    data = list(row)
    email_val = row[email_col]
    dept = ''
    if email_val:
        dept = dept_lookup.get(str(email_val).strip().lower(), '')
        if dept:
            matched += 1
    data.append(dept)
    for ci, v in enumerate(data, 1):
        out_ws.cell(row=row_num, column=ci, value=v)
    row_num += 1

src_wb.close()

# Auto-width
for col in out_ws.columns:
    mx = max((len(str(c.value or '')) for c in col), default=10)
    out_ws.column_dimensions[col[0].column_letter].width = min(mx + 3, 40)

# Save
output = os.path.join(os.getcwd(), 'public', 'emails_combined.xlsx')
out_wb.save(output)
out_wb.close()

print(f"Total: {row_num - 2}")
print(f"With dept: {matched}")
print(f"Without dept: {row_num - 2 - matched}")
print(f"Saved: {output}")
